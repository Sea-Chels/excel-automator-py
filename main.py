import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):  # we start at 2 because the first row is just names not values
        cell = sheet.cell(row, 3)  # we can access a specific cell via coordinates (row, column)
        # correct the price
        correct_price = cell.value * 0.9
        # grab the following cell in the same row and set value to the corrected price
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = correct_price

    # grabs all values only in column 4 in all rows.
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

process_workbook('transactions.xlsx')